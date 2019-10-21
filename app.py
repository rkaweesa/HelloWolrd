"-------------------------------------------------"
"-------------------------------------------------"
#Project 2: Machine Learning with Python  (Subset of AI)
#1. Import Data (CSV File)
#2. Clean the Data eg remove duplicated data
#3. Split the Data into Training/ test sets
#4. Create a Model e.g Physic learn
#5. Train Model
#6. Make predictions
#7. Evaluate and improve




"-------------------------------------------------"
"-------------------------------------------------"
#Excel Spreadsheets
#Project 1: Automation with Python


import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_worbook(filename):
    wb= xl.load_workbook('transactions.xlsx')
    sheet = wb['Sheet1']

    #This adds new discounted colom
    for row in range(2, sheet.max_row +1):
        cell = sheet.cell(row, 3)
        discounted_price = cell.value*0.9
        discounted_price_cell = sheet.cell(row, 4)
        discounted_price_cell.value = discounted_price

    #add chart
    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart (chart,'e2')

    wb.save (filename)

#you can now call specific files in different directories


"-------------------------------------------------"
"-------------------------------------------------"
#Pypi and Pip
#These are python packages that can be reused
#use pypi.org

#Search Openpyxl: Used for working with exceel spreedsheets

"-------------------------------------------------"
"-------------------------------------------------"
#PYTHON MODULES
#path

from pathlib import Path

#Program to automate, to iterate all spreedsheets in directory open and process them
path = Path ()
for file in  (path.glob('*.py')):
    print(file)

#Absolute path- we start from root of hard disk
# c\Program Files \Microsoft

#Relative path -ie path starting from current directory

path = Path ("ecommerce")
print (path.exists())



"-------------------------------------------------"
"-------------------------------------------------"
#GENETRATING RANDOM VALUES
#roll a dice

import random


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second =  random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())

"-------------------------------------------------"
#list of team members and pick out one randomly as leader
import random
members = ['Kay', 'Lay', 'May', 'Day', 'Pay', 'Bray']
leader = random.choice(members)
print(leader)

"-------------------------------------------------"


import random

for i in range (5):
    print(random.randint(10, 20))


"-------------------------------------------------"
"-------------------------------------------------"
#PACKAGES
#These are basically another way to organize code
#Its a container for related multiple modules
# A package is a directory or folder

#to import the entire module
from ecommerce import shipping

shipping.calc_shipping()

#importing module from ecommerce package

from ecommerce.shipping import calc_shipping

import ecommerce.shipping
ecommerce.shipping.calc_shipping()


"-------------------------------------------------"
"-------------------------------------------------"
#MODULES
#This is a file with python code used to organize code into multiple files
#create a new project file called conveter.py and import code here
#Each file is called a module containing all the functions and classes

from utils import find_max

numbers =[4, 5, 6, 72, 40, 65, 1, 66, 10, 2]
maximum = find_max(numbers)
print(maximum)


import converters
from converters import kg_to_lbs

kg_to_lbs(100)

print (converters.kg_to_lbs(70))


"-------------------------------------------------"
"-------------------------------------------------"
#INHERITANCE
#This is mechanism for resuing code


class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self):
        print("bark")


class Cat(Mammal):
    def purrl(self):
        print("purrl")


dog1 = Dog()
dog1.walk()

dog2 = Dog()
dog2.bark()

cat1 = Cat()
cat1.purrl()


"-------------------------------------------------"
class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    pass


class Cat(Mammal):
    pass


dog1 = Dog()
dog1.walk()

"-------------------------------------------------"
"-------------------------------------------------"
#CONSTRUCTORS
#create a point object without an x and y cordinate
#This is function that gets called at a time of creating and object


class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"Hi I am {self.name}")


roger = Person("Roger Kaweesa")
roger.talk()

kim = Person("Kim K")
kim.talk()

#Here each object is a different instance of the person class

"-------------------------------------------------"
class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print("talk")


roger = Person("Roger Kaweesa")
print(roger.name)
roger.talk()




class Point:
    def __init__(self, x, y):
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point= Point(10, 20)
point.x = 30
print(point.x)
"-------------------------------------------------"
"-------------------------------------------------"
#CLASSES


class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point()
point1.x = 10
point1.y = 20
print(f"({point1.x}, {point1.y})")

point2 = Point()
point2.x = 1
point2.y = 2
print(f"({point2.x}, {point2.y})")

"-------------------------------------------------"
"-------------------------------------------------"
"Comments"
#How to write comments
#This is a multiple comment
#We use comments to explais whys and hows and not what
print("Sky is blue")


"-------------------------------------------------"
"-------------------------------------------------"
#Exceptions
#These help in handling Errors by using try- expect"

try:
    age = int(input('Age: '))
    income = 20000
    risk = income/age
    print(age)
except ZeroDivisionError:
    print('Cannot have age = 0.')
except ValueError:
    print('Invalid value')

"-------------------------------------------------"


try:
    age = int(input('Age: '))
    print(age)
except ValueError:
    print('Invalid value')
"-------------------------------------------------"
"-------------------------------------------------"
#Return Statements
#Emojis converter again


def emoji_converter(text):
    words = text.split(" ")
    emojis = {
        ":)": "😊",
        ":(": "☹"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output

text = input(">")
print(emoji_converter(text))

"-------------------------------------------------"


def square(number):
    return number*number


print(square(3))

"-------------------------------------------------"\
 "Keyword Arguments"


def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard')


print("Start")
greet_user("Roger", last_name="Kaweesa")
print("Finish")

'''these help increase readability of code
key word arguments should always come after positional
arguments'''


"-------------------------------------------------"
"Parameters"
'''How to pass information to user. By adding parameters
as place holders 
'''


def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard')


print("Start")
greet_user("Kaweesa","Roger")
print("Finish")


def greet_user(name):
    print(f'Hi {name}!')
    print('Welcome aboard')


print("Start")
greet_user("Roger")
greet_user("Emma")
print("Finish")


"Functions"
def greet_user():
    print('Hi User!')
    print('Welcome aboard')


print("Start")
greet_user()
print("Finish")

"Dictionaries"
"Emojis converter"
text = input(">")
words = text.split(' ')
emojis = {
    ":)":"😊",
    ":(":"☹"
}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)


"Excerise mapping digits"
phone = input("Phone: ")
digits_mapping ={
"1": "One",
"2": "Two",
"3": "Three",
"4": "Four",
"5": "Five"
}
output = ""
for ch in phone:
    output+= digits_mapping.get(ch, "!")
print(output)


customer = {
    "name": "Roger Kaweesa",
    "age": 35,
    "is_verified": True
}
customer["sex"] = "male"
print(customer["sex"])


customer = {
    "name": "Roger Kaweesa",
    "age": 35,
    "is_verified": True
}
customer["name"] = "Roger Bambino"
print(customer["name"])


customer = {
    "name": "Roger Kaweesa",
    "age": 35,
    "is_verified": True
}
print(customer.get("birthdate", "It won't work"))



"Unpacking "
'''coordinates = (1, 2, 3)
x= coordinates[0]
y= coordinates[1]
z= coordinates[2] this is the same as below unpacking feature'''

coordinates = (1, 2, 3)
x, y, z = coordinates
print(y)

"Tuples is a structure and is similar to lists and can be used to store but unlike lists you cannot modify them ie immutable"
numbers = (1, 2, 3)
print(numbers[0])

"List methods/ functions"
'how to remove duplicates in a list'
numbers = [2, 3, 1, 4, 5, 3, 7, 7, 8, 9, 3, 5, 2]
uniques = []
for number in numbers:
    if number not in uniques:
        uniques.append(number)

numbers = [4, 5, 3, 8, 6, 7, 5]
numbers2= numbers.copy()
numbers.append(11)
print(numbers2)

numbers = [4, 5, 3, 8, 6, 7, 5]
numbers.sort()
numbers.reverse()
print(numbers)

numbers = [4, 5, 3, 8, 6, 7, 5]
print(numbers.count(5))

numbers = [4, 5, 3, 8, 6, 7, 5]
print(50 in numbers)


numbers = [4, 5, 3, 8, 6, 7]
print(numbers.index(8))

numbers = [4, 5, 3, 8, 6, 7]
numbers.pop(0)
print(numbers)


numbers = [4, 5, 3, 8, 6, 7]
numbers.insert(0, 10)
print(numbers)

numbers = [4, 5, 3, 8, 6, 7]
numbers.remove(4)
print(numbers)


"2D lists Matrix"
'using nested loops in 2D lists'
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
for row in matrix:
    for item in row:
        print(item)


'3x3 Matrix'
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
matrix[0][1]=30
print(matrix[0][1])


'lists'
numbers =[4,5,6,72, 40, 65, 1, 66,10,2]
max = numbers [0]
for number in numbers:
    if number> max:
        max = number
print(max)


names = ['John', 'Roger', 'Tim', 'Guy', 'Jane']
print(names[0:4])

'for loops'
numbers = [2, 2, 2, 2, 5]
for x_count in numbers:
    output = ''
    for count in range(x_count):
        output += 'x'
    print(output)



for x in range (5):
    for y in range (4):
        print(f'({x}, {y})')



cart = [10, 20, 30, 100]

total = 0
for price in cart:
    total += price
print(f"Total: {total}")


for item in range(0,11,2):
    print(item)



for item in ['Roger', 'Simon', 'Peter']:
    print(item)

for item in 'Kaweesa':
    print(item)

car_command = ""
started = False

'while loop'
while True:
    car_command = input("> ").lower()
    if car_command == "start":
        if started:
            print("Car is already started")
        else:
            started = True
            print("Car Started---- Ready to go!")
    elif car_command == "stop":
        if not started:
            print("Car is already stopped")
        else:
            started = False
            print("Car Stopped.")
    elif car_command == "help":
        print('''
start - to start the car
stop - to stop the car
quit - to exit
        ''')
    elif car_command == "quit":
        break
    else:
        print("Sorry I don't understand that")


secret_no = 10
guess_count= 0
guess_limit= 3
while guess_count < guess_limit:
    guess= int(input('Guess:'))
    guess_count+=1
    if guess== secret_no:
        print('You won')
        break
else:
    print('Sorry you failed')



x= 2
while x <=10:
    print('*'*x)
    x = x+2
print("Done")


x= 2
while x <=10:
    print(x)
    x = x+2
print("Done")


weight = int(input('weight: '))
unit = input('(L)bs or (K)g: ')

if unit.upper() == "L":
    weight_lb = (float(weight)*0.45)
    print (f" You are {weight_lb} kgs")
elif unit.upper() == "K":
    weight_kg = (float(weight)/0.45)
    print (f"You are {weight_kg} pounds")
else:
    print("Please enter correct Unit 'K' for Kgs or 'L' for Pounds")



name = "Kaweesa"
if len(name) <3:
    print("name must be at least 3 characters long")
elif len(name) >50:
    print("name can be a maximum of 50 characters long")
else:
    print("Name looks good")




temp = 30
if temp >30:
    print("It's a hot day")
else:
    print("Its not a hot day")





has_high_income = True
has_good_credit = False
has_criminal_record= True

if has_high_income and not has_criminal_record:
    print("Eligible for loan")

price = int(1000000)
good_credit = True
bad_credit = False

if good_credit:
    down_payment = price * 0.1
else:
    down_payment = price * 0.2

print(f"Down Payment for good credit is ${down_payment}")




is_hot = False
is_cold = False


if is_hot:
    print("It's a hot day")
    print("Drink Plenty of water")
elif is_cold:
    print("It's a cold day")
    print("Wear something warm")
else:
    print("Nice day")



import math
print(math.ceil(2.7))


x= -5.52
print(abs(x))

x= 5.52
print(round(x))

x= (10 + 2) * 3 - 2**2
print(x)

x = 20
x= x/2
x /=2
print(x)

course = 'Python for Beginners'
print ('Python' in course)

course = 'Python for Beginners'
print (course.replace('for', 'with'))

course = 'Python for Beginners'
print (course.find('f'))

course = 'Python for Beginners'
print (course.upper())

course = 'Python for Beginners'
print(len(course))


first = 'Roger '
last = 'Kaweesa'
msg = f'{first} [{last}] is a great coder. This is a formatted string and is easy to visualize'
print(msg)

first = 'Roger '
last = 'Kaweesa'
full_name = first + '['+ last+'] is a coder. The above method is better'
print(full_name)


first = 'Roger '
last = 'Kaweesa '
full_name = first + last+ 'is a coder'
print(full_name)

name = 'Jennifer'
print(name[1:-2])


course = 'Python for Beginners'
another = course [:]
print(another)

course = 'Python for Beginners'
print(course[1:])

course = 'Python for Beginners'
print(course[0:3])


course = 'Python for Beginners'
print(course[-2])


course ='''
Hi John,
This is our first email to you.

Thanks

Support team

'''
print(course)


course = "Python's course for Beginners"
print(course)

course = 'Python for "Beginners"'
print(course)

birth_year = input('Birth year: ')
print(type(birth_year))
age = 2019 - int(birth_year)
print(type(age))
print(age)

weight_lb = input('What is your weight in pounds: ')
print(type(weight_lb))
kg_weight = 0.45 * float(weight_lb)
print(type(kg_weight))
print(kg_weight)







